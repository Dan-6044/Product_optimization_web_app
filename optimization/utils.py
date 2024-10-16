import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def generate_excel_dashboard(file_path):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path)

    # Perform some simple analysis (e.g., summing values in a column)
    summary = df.describe()  # Summary statistics
    analysis_result = summary.to_json()  # Save the result as JSON (or any format)

    # Open the Excel file with openpyxl to create visualizations
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Add a summary sheet
    summary_sheet = wb.create_sheet("Summary")
    for r in dataframe_to_rows(df.describe(), index=True, header=True):
        summary_sheet.append(r)

    # Create a BarChart as a visualization
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row, max_col=ws.max_column)
    chart.add_data(data, titles_from_data=True)
    chart.title = "Data Visualization"
    ws.add_chart(chart, "E5")  # Adjust the position to place the chart in the sheet

    # Save the workbook with the chart
    dashboard_file_path = file_path.replace('.xlsx', '_dashboard.xlsx')
    wb.save(dashboard_file_path)

    return analysis_result, dashboard_file_path
